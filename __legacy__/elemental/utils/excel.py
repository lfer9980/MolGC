"""
Original Code: Abimael Guzman Pando
Refactored: Angel Fernandez
File: ReadExcel.py
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import plotly.express as px

from elemental.utils.topsis import TopsisUW


class ReadExcel:
    """
    """

    def __init__(self, main_dir: str, output_dir: str):
        """

            :param main_dir:
            :param output_dir:
        """
        self.main_dir = main_dir
        self.output_dir = output_dir
        self.file_map = {}
        self.family_groups = []

    def __call__(self):
        """
        """
        self.discover_files()
        self.group_by_family()

        all_p1, all_p2, fams = [], [], []
        for fam, files in self.family_groups:
            t1, p1, p2, _ = self.process_family(fam, files)
            self.save_family_results(fam, t1, p1, p2)
            all_p1.append(p1)
            all_p2.append(p2)
            fams.append(fam)

        self.compile_global(all_p1, all_p2, fams)

    def discover_files(self):
        """
           self.file_map with paths to ‘Table_MAE_general.xlsx’.
        """
        for func in os.listdir(self.main_dir):
            fpath = os.path.join(self.main_dir, func)
            if not os.path.isdir(fpath):
                continue
            self.file_map[func] = {}
            for fam in os.listdir(fpath):
                fam_path = os.path.join(fpath, fam)
                excel_file = os.path.join(fam_path, 'Table_MAE_general.xlsx')
                if os.path.isfile(excel_file):
                    self.file_map[func][fam] = excel_file

    def group_by_family(self):
        """
             Group routes by family, assuming the same families in each functional.
        """
        families = set().union(*(m.keys() for m in self.file_map.values()))
        for fam in sorted(families):
            files = [self.file_map[func][fam]
                     for func in sorted(self.file_map)
                     if fam in self.file_map[func]]
            if files:
                self.family_groups.append((fam, files))

    @classmethod
    def process_family(cls, family: str, files: list):
        """
            Reads and concatenates the excels of a family,
            calculates T1, P1, P2 (average MAE, normalized TOPSIS and ranking).
            :param: family;
            :param: files;
            :return:
        """

        dfs = [pd.read_excel(fp) for fp in files]
        data = pd.concat(dfs, ignore_index=True)

        last_col = data.columns[-1]
        idx = data[last_col]
        df = data.drop(columns=[data.columns[-1], data.columns[-2]])
        df.insert(0, last_col, idx)
        df = df.fillna(0)

        mean_tabla = df.groupby(df.columns[-1]).mean().T.reset_index()
        mean_tabla.columns = ['Functional'] + list(df.columns[-1:].unique())

        topsis = TopsisUW(mean_tabla)
        t_res, t_norm = topsis()

        p1 = mean_tabla.copy()
        p1['Average MAE'] = p1.iloc[:, 1:].mean(axis=1)
        p1 = p1.sort_values('Average MAE').reset_index(drop=True)
        p1.index += 1

        p2 = t_res[['Functional', 'Ranking']].sort_values('Ranking').reset_index(drop=True)
        p2.index += 1

        return mean_tabla, p1, p2, family

    def save_family_results(self, family: str, t1, p1, p2):
        """
            Save an Excel file with all tables and generate graphs.
            :param: family;
            :param: t1;
            :param: p1;
            :param: p2;
            :return:

        """
        out = os.path.join(self.output_dir, f'{family}_results')
        os.makedirs(out, exist_ok=True)
        excel_path = os.path.join(out, f'{family}_general_results.xlsx')

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            t1.to_excel(writer, sheet_name=family, index=False, startrow=2)
            p1.to_excel(writer, sheet_name=family, index=True, startcol=t1.shape[1] + 2, startrow=2)
            p2.to_excel(writer, sheet_name=family, index=True, startcol=t1.shape[1] + 2, startrow=p1.shape[0] + 6)

        wb = load_workbook(excel_path)
        ws = wb[family]
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 20
        wb.save(excel_path)

        fig1 = px.scatter(p1, x='Functional', y='Average MAE', text='Average MAE')
        fig1.update_traces(textposition='top center').update_layout(title=f'{family} MAE')
        fig1.write_image(os.path.join(out, f'{family}_MAE.png'))

        fig2 = px.scatter(p2, x='Functional', y='Ranking', text='Ranking')
        fig2.update_traces(textposition='top center').update_layout(title=f'{family} TOPSIS Ranking')
        fig2.write_image(os.path.join(out, f'{family}_TOPSIS.png'))

    def compile_global(self, all_p1: list, all_p2: list, families: list):
        """
            Builds and saves global tables and graphs.
            :param: all_p1;
            :param: all_p2;
            :param: families;
        """
        df_mae = pd.concat([p1.set_index('Functional')['Average MAE'] for p1 in all_p1], axis=1)
        df_mae.columns = families
        df_mae['Overall Average MAE'] = df_mae.mean(axis=1)
        df_mae = df_mae.sort_values('Overall Average MAE')

        df_rank = pd.concat([p2.set_index('Functional')['Ranking'] for p2 in all_p2], axis=1)
        df_rank.columns = families
        df_rank['Overall Avg Rank'] = df_rank.mean(axis=1)
        df_rank = df_rank.sort_values('Overall Avg Rank')

        out = os.path.join(self.output_dir, 'Overall_results')
        os.makedirs(out, exist_ok=True)
        with pd.ExcelWriter(os.path.join(out, 'Overall_results.xlsx')) as writer:
            df_mae.to_excel(writer, sheet_name='MAE')
            df_rank.to_excel(writer, sheet_name='TOPSIS')

        fig_mae = px.scatter(df_mae.reset_index(), x='Functional', y='Overall Average MAE', text='Overall Average MAE')
        fig_mae.update_traces(textposition='top center').write_image(os.path.join(out, 'Overall_MAE.png'))

        fig_rank = px.scatter(df_rank.reset_index(), x='Functional', y='Overall Avg Rank', text='Overall Avg Rank')
        fig_rank.update_traces(textposition='top center').write_image(os.path.join(out, 'Overall_TOPSIS.png'))


if __name__ == "__main__":
    processor = ReadExcel(
        main_dir="../Files",
        output_dir="../CrossValidationResults"
    )
    processor()
